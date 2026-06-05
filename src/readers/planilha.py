"""
Leitor da planilha Excel de Receitas e Despesas (Receitas_e_Despesas_2026_X_2.xlsx).

Layout: cada mês é uma aba.
  Linha 1: datas nas colunas pares (B, D, F, H...)
  Linha 2: "SANTA MARIA" — início das despesas SM
  Linhas 3..N: (descrição, valor) por coluna-par, SM
  Linha K: "Sub. Total" SM
  Linha K+1: "CANOAS" — início das despesas Canoas
  Linhas K+2..M: (descrição, valor) por coluna-par, Canoas
  Linha M+1: "Sub. Total" Canoas
  Linha M+2: "Total Saídas"
  Linha M+4: "Entr. SM"   — créditos SM por dia
  Linha M+5: "Entr. Canoas"
  Linha M+6: "SALDO"
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

import openpyxl
import pandas as pd


def _limpar_br(v) -> float:
    if v is None or (isinstance(v, float) and v != v):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    import re
    s = str(v).strip().replace("R$", "").replace("\xa0", "").replace(" ", "")
    s = re.sub(r"[^\d,.\-]", "", s)
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    elif "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def ler_planilha(
    caminho: Union[str, Path],
    meses: Optional[list[str]] = None,
) -> pd.DataFrame:
    """
    Lê a planilha mensal de despesas e retorna DataFrame padronizado.
    meses: lista de nomes de abas a ler. None = todas as abas de meses.
    """
    caminho = Path(caminho)
    wb = openpyxl.load_workbook(caminho, data_only=True)

    abas_meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
    ]
    if meses:
        abas_para_ler = [m for m in meses if m in wb.sheetnames]
    else:
        abas_para_ler = [a for a in abas_meses if a in wb.sheetnames]

    frames = []
    for nome_aba in abas_para_ler:
        ws = wb[nome_aba]
        df = _ler_aba(ws, nome_aba)
        if not df.empty:
            frames.append(df)

    if not frames:
        return pd.DataFrame(columns=["data", "banco", "descricao", "credito", "debito", "saldo", "horario", "documento", "unidade"])

    return pd.concat(frames, ignore_index=True)


def _ler_aba(ws, nome_aba: str) -> pd.DataFrame:
    """Lê uma aba mensal e retorna registros normalizados."""
    # Linha 1: datas → monta índice coluna → data
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_para_data: dict[int, datetime] = {}
    for col_idx, v in enumerate(row1):
        if isinstance(v, datetime):
            col_para_data[col_idx] = v  # col_idx 0-based

    if not col_para_data:
        return pd.DataFrame()

    # Detecta onde começa CANOAS e onde terminam os dados
    limite_sm:     Optional[int] = None
    inicio_canoas: Optional[int] = None
    limite_canoas: Optional[int] = None
    linha_entr_sm:     Optional[int] = None
    linha_entr_canoas: Optional[int] = None

    todos_rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(todos_rows):
        v0 = str(row[0]).strip().upper() if row[0] else ""
        if "CANOAS" in v0 and inicio_canoas is None:
            inicio_canoas = i
        if "SUB. TOTAL" in v0 or "SUB TOTAL" in v0:
            if inicio_canoas is None:
                limite_sm = i
            else:
                limite_canoas = i
        if "ENTR. SM" in v0 or "ENTR SM" in v0:
            linha_entr_sm = i
        if "ENTR. CANOAS" in v0 or "ENTR CANOAS" in v0:
            linha_entr_canoas = i

    registros: list[dict] = []

    def _extrair_despesas(inicio_linha: int, fim_linha: int, unidade: str) -> None:
        for row in todos_rows[inicio_linha:fim_linha]:
            for col_data_idx, data in col_para_data.items():
                # Descrição na col_data_idx, valor na col_data_idx + 1
                desc_col = col_data_idx      # 0-based
                val_col  = col_data_idx + 1
                if val_col >= len(row):
                    continue
                desc = row[desc_col]
                val  = row[val_col]
                if not desc or not val:
                    continue
                v = _limpar_br(val)
                if v == 0.0:
                    continue
                registros.append({
                    "data":      data,
                    "banco":     "Planilha",
                    "unidade":   unidade,
                    "descricao": str(desc).strip(),
                    "credito":   0.0,
                    "debito":    v,
                    "saldo":     0.0,
                    "horario":   None,
                    "documento": "",
                    "fonte":     "planilha",
                })

    def _extrair_entradas(linha_idx: int, unidade: str) -> None:
        if linha_idx is None or linha_idx >= len(todos_rows):
            return
        row = todos_rows[linha_idx]
        for col_data_idx, data in col_para_data.items():
            val_col = col_data_idx + 1
            if val_col >= len(row):
                continue
            val = row[val_col]
            if not val:
                continue
            v = _limpar_br(val)
            if v == 0.0:
                continue
            registros.append({
                "data":      data,
                "banco":     "Planilha",
                "unidade":   unidade,
                "descricao": f"ENTRADA {unidade}",
                "credito":   v,
                "debito":    0.0,
                "saldo":     0.0,
                "horario":   None,
                "documento": "",
                "fonte":     "planilha",
            })

    # Santa Maria: linhas 2 (índice 1) até limite_sm (exclusive)
    inicio_sm = 2  # linha 3 (0-indexed = 2), pula linha 1 (datas) e linha 2 (label)
    fim_sm = limite_sm if limite_sm else (inicio_canoas if inicio_canoas else len(todos_rows))
    _extrair_despesas(inicio_sm, fim_sm, "Santa Maria")

    # Canoas
    if inicio_canoas:
        fim_canoas = limite_canoas if limite_canoas else len(todos_rows)
        _extrair_despesas(inicio_canoas + 1, fim_canoas, "Canoas")

    # Entradas
    _extrair_entradas(linha_entr_sm, "Santa Maria")
    _extrair_entradas(linha_entr_canoas, "Canoas")

    if not registros:
        return pd.DataFrame()

    df = pd.DataFrame(registros)
    df["data"] = pd.to_datetime(df["data"], errors="coerce")
    df["descricao"] = df["descricao"].astype(str)
    return df
