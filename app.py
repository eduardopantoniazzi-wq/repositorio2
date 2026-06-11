"""
Controle de Débito — Moinho de Trigo v2
Tabela: Débitos Previstos (planilha FC) × Débitos Efetivados (extratos)
"""

import io, sys, tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

sys.path.insert(0, str(Path(__file__).parent))

st.set_page_config(page_title="Controle de Débito — Moinho de Trigo",
                   page_icon="🌾", layout="wide")

_MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
          "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
_MESES_NUM = {m: f"{i+1:02d}" for i, m in enumerate(_MESES)}

BANCO_POR_NOME = {
    "bradesco_alimentos": "Bradesco Alimentos", "bradescalimentos": "Bradesco Alimentos",
    "bradescoalimentos": "Bradesco Alimentos", "bradesco alimentos": "Bradesco Alimentos",
    "bradesco": "Bradesco", "sicredi": "Sicredi",
    "banrisul": "Banrisul",
    "bb_alimentos": "BB Alimentos", "bbalimentos": "BB Alimentos",
    "bb": "BB",
}

# Palavras que identificam o arquivo de Consulta Operações do Banrisul
_CONSULTA_BANRISUL = ("banrisul_consulta", "consulta_banrisul",
                      "banrisul consulta", "consulta banrisul",
                      "banrisul_operacoes", "operacoes_banrisul",
                      "banrisul operacoes", "operacoes banrisul",
                      "consulta_operacoes", "consulta operacoes")

def eh_consulta_banrisul(nome: str) -> bool:
    n = nome.lower()
    # Qualquer arquivo com "banrisul" E "consulta" no nome
    if "banrisul" in n and "consulta" in n:
        return True
    return any(p in n for p in _CONSULTA_BANRISUL)

def detectar_banco(nome):
    n = nome.lower()
    for k, v in BANCO_POR_NOME.items():
        if k in n:
            return v
    return None

def ler_extrato(nome, conteudo):
    from src.readers.bradesco import LeitorBradesco
    from src.readers.sicredi  import LeitorSicredi
    from src.readers.bb       import LeitorBB
    from src.readers.banrisul import LeitorBanrisul
    LEITORES = {"Bradesco": LeitorBradesco, "Bradesco Alimentos": LeitorBradesco,
                "Sicredi": LeitorSicredi,
                "BB": LeitorBB, "BB Alimentos": LeitorBB, "Banrisul": LeitorBanrisul}
    banco = detectar_banco(nome)
    if not banco:
        return None, f"'{nome}' não reconhecido — renomeie com bradesco/bradesco_alimentos/sicredi/bb/bb_alimentos/banrisul no início"
    suffix = Path(nome).suffix
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(conteudo); tmp_path = Path(tmp.name)
    try:
        df = LEITORES[banco](tmp_path).ler()
        df["banco"] = banco
        return df, None
    except Exception as e:
        return None, f"Erro ao ler '{nome}': {e}"
    finally:
        tmp_path.unlink(missing_ok=True)

def ler_planilha(conteudo, meses, filename=".xlsx"):
    from src.readers.planilha import ler_planilha as _ler
    suffix = Path(filename).suffix or ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(conteudo); tmp_path = Path(tmp.name)
    try:
        return _ler(tmp_path, meses=meses), None
    except Exception as e:
        return None, f"Erro ao ler planilha: {e}"
    finally:
        tmp_path.unlink(missing_ok=True)


def _mes_nome(numero: int) -> str:
    return _MESES[numero - 1]

def conciliar(df_prev, df_banco, limite_alerta: float = 1_500.0):
    """
    Casamento global: monta a matriz de scores de todos os pares possíveis
    e atribui o melhor par disponível globalmente (não guloso por linha).
    Isso evita que dois previstos com valor parecido troquem de beneficiário.
    """
    from difflib import SequenceMatcher
    import re

    STOP = {"nf","nota","fiscal","ltda","sa","eireli","me","epp","pag","pgto",
            "ted","pix","de","boleto","pagamento","transferencia","transf"}

    _norm_cache = {}
    def norm(s):
        if s not in _norm_cache:
            t = re.sub(r"[^\w\s]", " ", str(s).upper())
            _norm_cache[s] = " ".join(w for w in t.split() if len(w) > 2 and w.lower() not in STOP)
        return _norm_cache[s]

    def sim_nome(a, b):
        na, nb = norm(a), norm(b)
        if not na or not nb:
            return 0.0
        palavras_a = set(na.split())
        palavras_b = set(nb.split())
        comuns = palavras_a & palavras_b
        if comuns:
            return 0.55 + 0.1 * min(len(comuns), 3)
        return SequenceMatcher(None, na, nb).ratio()

    # Filtro vetorizado (mais rápido que .apply)
    _OP_EXCLUIR = ["TARIFA","TAXA","IOF","JUROS","INSS","FGTS","SALDO","RENTAB",
                   "FACILCRED","RENDE FACIL","DEBITO SERV"]
    _op_pat = "|".join(_OP_EXCLUIR)

    deb_banco = df_banco[
        (df_banco["debito"] > 0) &
        ~df_banco["descricao"].str.upper().str.contains(_op_pat, regex=True, na=False)
    ].copy().reset_index(drop=True)
    deb_prev  = df_prev[df_prev["debito"] > 0].copy().reset_index(drop=True)

    # Pré-calcula normas e palavras para evitar recalcular no loop
    prev_norms = {ip: (norm(r["descricao"]), set(norm(r["descricao"]).split()))
                  for ip, r in deb_prev.iterrows()}
    banco_norms = {ib: (norm(r["descricao"]), set(norm(r["descricao"]).split()))
                   for ib, r in deb_banco.iterrows()}

    nP = len(deb_prev)
    nB = len(deb_banco)

    # ── Monta matriz de scores (nP × nB) ────────────────────────────────
    scores = {}   # (ip, ib) -> score
    for ip, prev in deb_prev.iterrows():
        na, palavras_a = prev_norms[ip]
        for ib, deb in deb_banco.iterrows():
            diff_val = abs(deb["debito"] - prev["debito"]) / max(prev["debito"], 1)
            if diff_val > 0.60:
                continue
            try:
                diff_dias = abs((deb["data"] - prev["data"]).days)
            except Exception:
                diff_dias = 999
            if diff_dias > 5:
                continue

            nb, palavras_b = banco_norms[ib]
            comuns = palavras_a & palavras_b

            # Casamento 1:1 exige pelo menos uma palavra em comum — sem isso, não casa
            if not comuns:
                continue

            s_nome = 0.55 + 0.1 * min(len(comuns), 3)
            s_val  = max(0.0, 1 - diff_val / 0.60)
            s_data = max(0.0, 1 - diff_dias / 6)

            # Nome tem peso maior que valor — evita trocar beneficiários
            score = s_nome * 0.50 + s_val * 0.35 + s_data * 0.15

            scores[(ip, ib)] = score

    # ── Casamento global: atribui pares por ordem de melhor score ────────
    pares_ordenados = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    usados_prev  = set()
    usados_banco = set()
    atribuicoes  = {}   # ip -> (ib, score)

    for (ip, ib), score in pares_ordenados:
        if ip in usados_prev or ib in usados_banco:
            continue
        if score < 0.30:
            break
        atribuicoes[ip] = (ib, score)
        usados_prev.add(ip)
        usados_banco.add(ib)

    # ── Casamento 1:N — um previsto para múltiplos boletos ───────────────
    # Coleta todos os débitos bancários livres com suas palavras normalizadas
    livres_info = [
        (ib, deb_banco.loc[ib, "debito"], set(norm(deb_banco.loc[ib, "descricao"]).split()))
        for ib in deb_banco.index
        if ib not in usados_banco
    ]

    atribuicoes_multi = {}

    for ip, prev in deb_prev.iterrows():
        if ip in atribuicoes:
            continue
        prev_val   = prev["debito"]
        prev_norm  = norm(prev["descricao"])
        palavras_p = set(prev_norm.split())
        if not palavras_p:
            continue

        # Candidatos: boletos livres menores que o total, com pelo menos 1 palavra em comum
        cands = [(ib, v) for ib, v, wset in livres_info
                 if ib not in usados_banco and v < prev_val and (palavras_p & wset)]

        if len(cands) < 2:
            continue
        if sum(v for _, v in cands) < prev_val * 0.75:
            continue

        # Greedy: maiores valores primeiro, acumula até ≈ alvo (tolerância 15%)
        cands.sort(key=lambda x: x[1], reverse=True)
        sel, total = [], 0.0
        for ib, v in cands:
            if total + v <= prev_val * 1.15:
                sel.append(ib)
                total += v

        if len(sel) >= 2 and abs(total - prev_val) / max(prev_val, 1) <= 0.15:
            atribuicoes_multi[ip] = (sel, total)
            for ib in sel:
                usados_banco.add(ib)
            usados_prev.add(ip)

    LIMITE_ALERTA = float(limite_alerta)

    # ── Monta linhas da tabela ───────────────────────────────────────────
    linhas = []

    for ip, prev in deb_prev.iterrows():
        if ip in atribuicoes_multi:
            ibs, total = atribuicoes_multi[ip]
            debs     = deb_banco.loc[ibs]
            total    = round(total, 2)
            diff     = round(total - prev["debito"], 2)
            pct      = diff / prev["debito"] * 100 if prev["debito"] else 0
            alerta   = f"🔴 Diferença de R$ {abs(diff):,.2f}" if abs(diff) > 0.01 else ""
            status   = "✅ OK (múlt.)" if abs(diff) <= prev["debito"] * 0.02 else "⚠️ VALOR DIFERENTE (múlt.)"
            banco_str = debs["banco"].iloc[0] if debs["banco"].nunique() == 1 else "Múltiplos"
            benef     = debs["descricao"].iloc[0].split("/")[0].strip()
            linhas.append({
                "Status":                status,
                "🔴 Alerta":             alerta,
                "Data Prevista":         prev["data"],
                "Beneficiário Previsto": prev["descricao"],
                "Valor Previsto (R$)":   prev["debito"],
                "Data Pago":             debs["data"].min(),
                "Banco":                 banco_str,
                "Pago Para":             f"Múltiplos boletos ({len(ibs)}x) / {benef}",
                "Valor Pago (R$)":       total,
                "Diferença (R$)":        diff,
                "Diferença (%)":         round(pct, 1),
            })
            continue

        if ip in atribuicoes:
            ib, score = atribuicoes[ip]
            deb  = deb_banco.loc[ib]
            diff = round(deb["debito"] - prev["debito"], 2)
            pct  = diff / prev["debito"] * 100 if prev["debito"] else 0
            s_nome = sim_nome(prev["descricao"], deb["descricao"])

            if abs(diff) <= prev["debito"] * 0.02 and s_nome >= 0.40:
                status = "✅ OK"
            elif abs(diff) > prev["debito"] * 0.02 and s_nome >= 0.40:
                status = "⚠️ VALOR DIFERENTE"
            elif abs(diff) <= prev["debito"] * 0.02 and s_nome < 0.40:
                status = "⚠️ BENEFICIÁRIO DIFERENTE"
            else:
                status = "⚠️ DIVERGÊNCIA"

            # Alerta somente quando há diferença de valor (independente do beneficiário)
            if abs(diff) > 0.01:
                alerta = f"🔴 Diferença de R$ {abs(diff):,.2f}"
            else:
                alerta = ""

            linhas.append({
                "Status":                  status,
                "🔴 Alerta":               alerta,
                "Data Prevista":           prev["data"],
                "Beneficiário Previsto":   prev["descricao"],
                "Valor Previsto (R$)":     prev["debito"],
                "Data Pago":               deb["data"],
                "Banco":                   deb["banco"],
                "Pago Para":               deb["descricao"],
                "Valor Pago (R$)":         deb["debito"],
                "Diferença (R$)":          diff,
                "Diferença (%)":           round(pct, 1),
            })
        else:
            linhas.append({
                "Status":                  "🕐 NÃO PAGO",
                "🔴 Alerta":               "",
                "Data Prevista":           prev["data"],
                "Beneficiário Previsto":   prev["descricao"],
                "Valor Previsto (R$)":     prev["debito"],
                "Data Pago":               None,
                "Banco":                   "",
                "Pago Para":               "",
                "Valor Pago (R$)":         None,
                "Diferença (R$)":          -prev["debito"],
                "Diferença (%)":           -100.0,
            })

    # ── Débitos bancários sem par = não previstos ────────────────────────
    OPERACIONAL = ["TARIFA","TAXA","IOF","JUROS","INSS","FGTS","SALDO","RENTAB",
                   "FACILCRED","RENDE FACIL","DEBITO SERV"]
    for ib, deb in deb_banco.iterrows():
        if ib in usados_banco:
            continue
        desc = str(deb["descricao"]).upper()
        if any(p in desc for p in OPERACIONAL):
            continue
        linhas.append({
            "Status":                  "🚨 NÃO PREVISTO",
            "🔴 Alerta":               "",
            "Data Prevista":           None,
            "Beneficiário Previsto":   "",
            "Valor Previsto (R$)":     None,
            "Data Pago":               deb["data"],
            "Banco":                   deb["banco"],
            "Pago Para":               deb["descricao"],
            "Valor Pago (R$)":         deb["debito"],
            "Diferença (R$)":          deb["debito"],
            "Diferença (%)":           None,
        })

    return pd.DataFrame(linhas)



def cor_linha(row):
    s = str(row.get("Status",""))
    if s.startswith("✅"):
        bg = "#d4edda; color:#155724"
    elif "BENEFICIÁRIO" in s:
        bg = "#f8d7da; color:#721c24"   # vermelho — nome trocado é mais grave
    elif s.startswith("⚠️"):
        bg = "#fff3cd; color:#856404"   # amarelo — só valor diferente
    elif s.startswith("🚨"):
        bg = "#f8d7da; color:#721c24"
    else:
        bg = "#e2e3e5; color:#383d41"
    return [f"background-color:{bg}"] * len(row)


def fmt_brl(v):
    try:    return f"R$ {float(v):,.2f}"
    except: return ""

def fmt_dt(v):
    try:    return pd.Timestamp(v).strftime("%d/%m/%Y")
    except: return ""

# ─────────────────────────────────────────────────────────────────────────────
st.title("🌾 Débitos Previstos × Efetivados")
st.caption("Moinho de Trigo — comparação diária: o que estava previsto na planilha vs o que saiu dos bancos")

# Sidebar
with st.sidebar:
    st.header("📂 Arquivos")
    extratos_up = st.file_uploader("Extratos bancários (PDF, XLSX, CSV)",
        type=["pdf","xlsx","xls","csv"], accept_multiple_files=True)
    planilha_up = st.file_uploader("Planilha de previsões (FC)",
        type=["xlsx","xls"])
    st.caption("Extratos: bradesco_*.pdf · bradesco_alimentos_*.pdf · sicredi_*.pdf · bb_*.pdf · bb_alimentos_*.pdf · banrisul_*.pdf")
    st.caption("Consulta Banrisul: banrisul_consulta_*.pdf (enriquece PGTO BOLETO e transferências com nome do beneficiário)")
    st.divider()
    data_sel = st.date_input("📅 Data do extrato", value=datetime.now().date())
    mostrar_periodo = st.toggle("Ver período (mais de um dia)", value=False)
    if mostrar_periodo:
        data_fim = st.date_input("Até", value=datetime.now().date())
    else:
        data_fim = data_sel
    rodar = st.button("▶ Comparar", type="primary", use_container_width=True)

if not rodar:
    st.markdown("""
### Como funciona
1. Faça upload dos **extratos bancários** do dia (PDF ou Excel)
2. Faça upload da **planilha de previsões** (FC)
3. Selecione o **mês** e clique em **▶ Comparar**

A tabela mostra linha por linha:

| Cor | O que significa |
|---|---|
| 🟢 Verde | Pago conforme previsto |
| 🟡 Amarelo | Pago, mas valor diferente do previsto |
| 🔴 Vermelho | **Saiu do banco sem estar na planilha** |
| ⬜ Cinza | Previsto mas ainda não pago |
""")
    st.stop()

if not extratos_up:
    st.error("Faça upload dos extratos bancários.")
    st.stop()
if not planilha_up:
    st.error("Faça upload da planilha de previsões.")
    st.stop()

# Lê extratos e consultas de operações
with st.spinner("Lendo extratos..."):
    dfs = []
    dfs_consulta = []   # arquivos de Consulta Operações do Banrisul
    for f in extratos_up:
        conteudo = f.read()
        if eh_consulta_banrisul(f.name):
            from src.readers.banrisul_consulta import ler_consulta_banrisul
            with tempfile.NamedTemporaryFile(suffix=Path(f.name).suffix, delete=False) as tmp:
                tmp.write(conteudo); tmp_path = Path(tmp.name)
            try:
                df_c = ler_consulta_banrisul(tmp_path)
                if not df_c.empty:
                    dfs_consulta.append(df_c)
                    st.sidebar.success(f"✔ {f.name} ({len(df_c)} operações — consulta Banrisul)")
            except Exception as e:
                st.warning(f"Erro ao ler consulta Banrisul '{f.name}': {e}")
            finally:
                tmp_path.unlink(missing_ok=True)
        else:
            df, err = ler_extrato(f.name, conteudo)
            if err:
                st.warning(err)
            else:
                dfs.append(df)
                st.sidebar.success(f"✔ {f.name} → **{df['banco'].iloc[0]}** ({len(df)} lançamentos)")

if not dfs:
    st.error("Nenhum extrato lido. Verifique os nomes dos arquivos.")
    st.stop()

df_banco = pd.concat(dfs, ignore_index=True)

# ── Enriquece descrições do Banrisul com nomes da Consulta Operações ────────
if dfs_consulta:
    df_cons = pd.concat(dfs_consulta, ignore_index=True)
    # Lookup: valor_arredondado → lista de (data, beneficiario)
    # Indexado só por valor para tolerar datas com 1 dia de diferença
    _lookup_val: dict = {}
    for _, row in df_cons.iterrows():
        v = round(row["valor"], 2)
        _lookup_val.setdefault(v, []).append((row["data"], row["beneficiario"]))

    _GENERICOS_BNR = {"pgto boleto", "pag boleto", "pagamento boleto",
                      "debito automatico", "arrecadacao", "cobranca",
                      "debito transferencia", "deb transferencia",
                      "transferencia", "ted", "pix"}

    def _enriquecer_banrisul(row):
        if row["banco"] != "Banrisul":
            return row["descricao"]
        desc_low = str(row["descricao"]).lower()
        if not any(g in desc_low for g in _GENERICOS_BNR):
            return row["descricao"]
        valor = round(float(row["debito"]), 2)
        if valor == 0:
            return row["descricao"]
        candidatos = _lookup_val.get(valor, [])
        if not candidatos:
            return row["descricao"]
        data_row = row["data"]
        # Tenta match exato de data primeiro, depois aceita ±1 dia
        for delta in (0, 1, -1):
            alvo = data_row + pd.Timedelta(days=delta)
            for i, (data_c, nome) in enumerate(candidatos):
                if data_c == alvo:
                    candidatos.pop(i)
                    return f"{row['descricao']} / {nome}"
        return row["descricao"]

    df_banco["descricao"] = df_banco.apply(_enriquecer_banrisul, axis=1)

df_banco_full = df_banco.copy()  # cópia completa para cálculo de saldo (antes do filtro de data)

# Determina meses necessários a partir das datas selecionadas
meses_necessarios = list({_mes_nome(d.month)
                          for d in pd.date_range(data_sel, data_fim, freq="MS").date.tolist()
                          or [data_sel]})
if not meses_necessarios:
    meses_necessarios = [_mes_nome(data_sel.month)]

# Lê planilha
with st.spinner("Lendo planilha de previsões..."):
    df_prev, err = ler_planilha(planilha_up.read(), meses=meses_necessarios, filename=planilha_up.name)
    if err:
        st.error(err); st.stop()
    st.sidebar.success(f"✔ {planilha_up.name} ({len(df_prev)} lançamentos)")

# Filtra pelo intervalo de datas
data_ini_ts = pd.Timestamp(data_sel)
data_fim_ts = pd.Timestamp(data_fim)

df_banco = df_banco[(df_banco["data"] >= data_ini_ts) & (df_banco["data"] <= data_fim_ts)].copy()
df_prev  = df_prev[(df_prev["data"]  >= data_ini_ts) & (df_prev["data"]  <= data_fim_ts)].copy()

if df_banco.empty and df_prev.empty:
    st.warning(f"Nenhum lançamento encontrado para {data_sel.strftime('%d/%m/%Y')}. "
               "Verifique se a data bate com os extratos enviados.")
    st.stop()

periodo_label = (data_sel.strftime("%d/%m/%Y") if data_sel == data_fim
                 else f"{data_sel.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}")
st.info(f"📅 Mostrando: **{periodo_label}** — "
        f"{len(df_banco[df_banco['debito']>0])} débitos bancários · "
        f"{len(df_prev[df_prev['debito']>0])} previstos")

# Saldos — usa dataset completo (sem filtro de data) para pegar o último saldo real do arquivo
st.subheader("💰 Saldos dos Bancos")
bancos_presentes = [b for b in ["Bradesco","Bradesco Alimentos","Sicredi","BB","BB Alimentos","Banrisul"]
                    if not df_banco_full[df_banco_full["banco"] == b].empty]
cols = st.columns(len(bancos_presentes) + 1)
for i, b in enumerate(bancos_presentes):
    sub_b = df_banco_full[df_banco_full["banco"] == b]
    sub_nz = sub_b["saldo"][sub_b["saldo"] != 0]
    s = float(sub_nz.iloc[-1]) if not sub_nz.empty else 0.0
    cols[i].metric(b, fmt_brl(s))

# Saldo FC da planilha para o dia selecionado
from src.readers.planilha import ler_saldo_fc as _ler_saldo_fc
saldo_fc = None
if planilha_up:
    try:
        import tempfile as _tmp
        with _tmp.NamedTemporaryFile(suffix=".xlsx", delete=False) as _t:
            planilha_up.seek(0); _t.write(planilha_up.read()); _tp = Path(_t.name)
        saldo_fc = _ler_saldo_fc(_tp, pd.Timestamp(data_sel))
        _tp.unlink(missing_ok=True)
    except Exception:
        pass

cols[-1].metric("📊 Saldo FC do Dia", fmt_brl(saldo_fc) if saldo_fc is not None else "—")

# Concilia
with st.spinner("Comparando previstos × efetivados..."):
    df_res = conciliar(df_prev, df_banco)

# Métricas
st.subheader("📊 Resumo")
n_ok   = df_res["Status"].str.startswith("✅").sum()
n_div  = df_res["Status"].str.startswith("⚠️").sum()
n_nao  = df_res["Status"].str.startswith("🚨").sum()
n_pend = df_res["Status"].str.startswith("🕐").sum()
v_div  = df_res[df_res["Status"].str.startswith("⚠️")]["Diferença (R$)"].abs().sum()
v_nao  = df_res[df_res["Status"].str.startswith("🚨")]["Valor Pago (R$)"].sum()

c1,c2,c3,c4 = st.columns(4)
c1.metric("✅ Conforme previsto", n_ok)
c2.metric("⚠️ Valor diferente",  n_div, delta=fmt_brl(v_div) if n_div else None, delta_color="inverse")
c3.metric("🚨 Fora da planilha", n_nao, delta=fmt_brl(v_nao) if n_nao else None, delta_color="inverse")
c4.metric("🕐 Ainda não pago",   n_pend)

# Filtros
st.subheader("📋 Tabela Comparativa")
cf1, cf2, cf3 = st.columns(3)
status_opts = sorted(df_res["Status"].unique())
status_sel  = cf1.multiselect("Filtrar por status", status_opts, default=status_opts)
banco_opts  = ["Todos"] + sorted(df_banco["banco"].unique())
banco_sel   = cf2.selectbox("Banco", banco_opts)
busca       = cf3.text_input("Buscar nome", placeholder="ex: Cooperativa, Fornecedor...")

df_view = df_res[df_res["Status"].isin(status_sel)].copy()
if banco_sel != "Todos":
    df_view = df_view[df_view["Banco"] == banco_sel]
if busca:
    mask = (df_view["Beneficiário Previsto"].str.contains(busca, case=False, na=False) |
            df_view["Pago Para"].str.contains(busca, case=False, na=False))
    df_view = df_view[mask]

# Formata para exibição
df_show = df_view.copy()
df_show["Data Prevista"]  = df_show["Data Prevista"].apply(fmt_dt)
df_show["Data Pago"]      = df_show["Data Pago"].apply(fmt_dt)
df_show["Valor Previsto (R$)"] = df_show["Valor Previsto (R$)"].apply(fmt_brl)
df_show["Valor Pago (R$)"]     = df_show["Valor Pago (R$)"].apply(fmt_brl)
df_show["Diferença (R$)"] = df_show["Diferença (R$)"].apply(
    lambda v: f"+{fmt_brl(v)}" if isinstance(v,(int,float)) and v>0 else
              (fmt_brl(v) if isinstance(v,(int,float)) else ""))
df_show["Diferença (%)"]  = df_show["Diferença (%)"].apply(
    lambda v: f"{v:+.1f}%" if isinstance(v,(int,float)) else "")

st.dataframe(
    df_show.style.apply(cor_linha, axis=1),
    use_container_width=True,
    height=620,
    column_config={
        "Status":                  st.column_config.TextColumn(width=200),
        "🔴 Alerta":               st.column_config.TextColumn(width=280),
        "Data Prevista":           st.column_config.TextColumn("Data Prev.", width=100),
        "Beneficiário Previsto":   st.column_config.TextColumn("Previsto Para", width=230),
        "Valor Previsto (R$)":     st.column_config.TextColumn("Vlr Previsto", width=130),
        "Data Pago":               st.column_config.TextColumn("Data Pago", width=100),
        "Banco":                   st.column_config.TextColumn(width=90),
        "Pago Para":               st.column_config.TextColumn(width=230),
        "Valor Pago (R$)":         st.column_config.TextColumn("Vlr Pago", width=130),
        "Diferença (R$)":          st.column_config.TextColumn("Diferença R$", width=120),
        "Diferença (%)":           st.column_config.TextColumn("Diferença %", width=100),
    },
)

# Download com formatação Excel
def _formatar_excel(wb):
    """Aplica formatação (cores, larguras, cabeçalho) a todas as abas do workbook."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Mapeamento status → cor de fundo (hex sem #)
    _COR_STATUS = {
        "✅": ("D4EDDA", "155724"),   # verde
        "⚠️": ("FFF3CD", "856404"),   # amarelo
        "🚨": ("F8D7DA", "721C24"),   # vermelho
        "🕐": ("E2E3E5", "383D41"),   # cinza
    }
    _COR_BENEF = ("F8D7DA", "721C24")  # vermelho para beneficiário trocado

    # Larguras preferidas por nome de coluna (em caracteres)
    _LARGURAS = {
        "Status": 32, "🔴 Alerta": 40,
        "Data Prevista": 14, "Data Pago": 14,
        "Beneficiário Previsto": 34, "Pago Para": 34,
        "Valor Previsto (R$)": 18, "Valor Pago (R$)": 18,
        "Diferença (R$)": 18, "Diferença (%)": 14,
        "Banco": 14,
    }

    borda_fina = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for ws in wb.worksheets:
        # ── Cabeçalho ──────────────────────────────────────────────────────
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor="2C3E50")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = borda_fina
        ws.row_dimensions[1].height = 30

        # Descobre índice da coluna Status
        status_col = None
        headers = [c.value for c in ws[1]]
        if "Status" in headers:
            status_col = headers.index("Status")

        # ── Linhas de dados ─────────────────────────────────────────────────
        for row in ws.iter_rows(min_row=2):
            # Determina cor pela coluna Status
            bg, fg = "FFFFFF", "000000"
            if status_col is not None:
                status_val = str(row[status_col].value or "")
                if "BENEFICIÁRIO" in status_val:
                    bg, fg = _COR_BENEF
                else:
                    for prefix, (b, f) in _COR_STATUS.items():
                        if status_val.startswith(prefix):
                            bg, fg = b, f
                            break

            fill = PatternFill("solid", fgColor=bg)
            font_cor = Font(color=fg, size=10)

            for cell in row:
                cell.fill      = fill
                cell.font      = font_cor
                cell.border    = borda_fina
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                col_name = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
                # Converte colunas monetárias para número real → Excel consegue somar
                if col_name in ("Valor Previsto (R$)", "Valor Pago (R$)", "Diferença (R$)"):
                    raw = str(cell.value or "").replace("R$", "").replace("+", "").replace(",", "").strip()
                    try:
                        cell.value = float(raw)
                        cell.number_format = '"R$ "#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    except Exception:
                        pass

        # ── Larguras das colunas ────────────────────────────────────────────
        for i, col_name in enumerate(headers, start=1):
            larg = _LARGURAS.get(col_name, 18)
            ws.column_dimensions[get_column_letter(i)].width = larg

        # Congela cabeçalho
        ws.freeze_panes = "A2"

    return wb


buf = io.BytesIO()
with pd.ExcelWriter(buf, engine="openpyxl") as w:
    df_show.to_excel(w, sheet_name="Comparativo", index=False)
    df_res[df_res["Status"].str.startswith("🚨")].to_excel(w, sheet_name="Fora da Planilha", index=False)
    df_res[df_res["Status"].str.startswith("⚠️")].to_excel(w, sheet_name="Divergências de Valor", index=False)
    df_banco.to_excel(w, sheet_name="Extrato Consolidado", index=False)

import openpyxl as _openpyxl
buf.seek(0)
_wb = _openpyxl.load_workbook(buf)
_formatar_excel(_wb)
buf2 = io.BytesIO()
_wb.save(buf2)

st.download_button("📥 Baixar Excel",
    data=buf2.getvalue(),
    file_name=f"comparativo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True)
